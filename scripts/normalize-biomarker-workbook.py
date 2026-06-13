from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPORT_SPECS = [
    {"label": "RLP-2018", "value_col": 11, "date_col": 12, "fallback_date": None, "date_note": ""},
    {"label": "RLP 2020", "value_col": 13, "date_col": 14, "fallback_date": None, "date_note": ""},
    {"label": "SYNLAB_IFMV2021", "value_col": 15, "date_col": 16, "fallback_date": "2021-01-05", "date_note": "fallback_report_date_used"},
    {"label": "SYNLAB_k 7.10.2021", "value_col": 17, "date_col": None, "fallback_date": "2021-10-07", "date_note": "date_inferred_from_header"},
    {"label": "RLP 2022", "value_col": 18, "date_col": None, "fallback_date": "2022-01-01", "date_note": "date_approximated_from_year_header"},
    {"label": "RLP 2024", "value_col": 19, "date_col": None, "fallback_date": "2024-01-01", "date_note": "date_approximated_from_year_header"},
    {"label": "RLP2026", "value_col": 20, "date_col": None, "fallback_date": "2026-01-01", "date_note": "date_approximated_from_year_header"},
]


MARKER_KEY_MAP = {
    "GLU": ("glucose_fasting", "Glukoza nalacno", "glucose_metabolism"),
    "GLUP": ("glucose_fasting", "Glukoza v plazme", "glucose_metabolism"),
    "HBA1C": ("hba1c", "HbA1c", "glucose_metabolism"),
    "GLHB": ("hba1c", "HbA1c", "glucose_metabolism"),
    "INS": ("insulin_fasting", "Inzulin nalacno", "glucose_metabolism"),
    "KRE": ("creatinine", "Kreatinin", "kidney"),
    "ALT": ("alt", "ALT", "liver"),
    "AST": ("ast", "AST", "liver"),
    "ALP": ("alp", "ALP", "liver"),
    "TBIL": ("bilirubin_total", "Bilirubin celkovy", "liver"),
    "DBIL": ("bilirubin_direct", "Bilirubin konjugovany", "liver"),
    "GGT": ("ggt", "GGT", "liver"),
    "CRP": ("crp", "CRP", "inflammation"),
    "HSCRP": ("hs_crp", "hs-CRP", "inflammation"),
    "HCRP": ("hs_crp", "hs-CRP", "inflammation"),
    "B12": ("vitamin_b12", "Vitamin B12", "vitamins"),
    "FOL": ("folate", "Folat", "vitamins"),
    "HCY": ("homocysteine", "Homocystein", "methylation"),
    "HCYSE": ("homocysteine", "Homocystein", "methylation"),
    "TSH": ("tsh", "TSH", "thyroid"),
    "FT3": ("ft3", "fT3", "thyroid"),
    "FT4": ("ft4", "fT4", "thyroid"),
    "ATPO": ("anti_tpo", "Anti TPO", "thyroid"),
    "ATG": ("anti_tg", "Anti TG", "thyroid"),
    "TRAK": ("trak", "TRAK", "thyroid"),
    "CHOL": ("total_cholesterol", "Celkovy cholesterol", "lipids"),
    "HDL": ("hdl_c", "HDL cholesterol", "lipids"),
    "LDL": ("ldl_c", "LDL cholesterol", "lipids"),
    "TRIG": ("triglycerides", "Triglyceridy", "lipids"),
    "TG": ("triglycerides", "Triglyceridy", "lipids"),
    "FERR": ("ferritin", "Ferritin", "minerals"),
    "FER": ("ferritin", "Ferritin", "minerals"),
    "25OHD": ("vitamin_d_25oh", "Vitamin D 25-OH", "vitamins"),
    "VITD": ("vitamin_d_25oh", "Vitamin D 25-OH", "vitamins"),
    "UREA": ("urea", "Urea", "kidney"),
    "DHEAS": ("dheas", "DHEAS", "hormones"),
    "KORT": ("cortisol", "Kortizol", "hormones"),
    "TST": ("testosterone_total", "Testosteron", "hormones"),
    "FE": ("iron_serum", "Zelezo", "minerals"),
    "TRANS": ("transferrin", "Transferin", "minerals"),
    "CA": ("calcium_ionized", "Vapnik ionizovany", "minerals"),
    "MG": ("magnesium", "Horcik", "minerals"),
    "NA": ("sodium", "Sodik", "electrolytes"),
    "K": ("potassium", "Draslik", "electrolytes"),
    "CL": ("chloride", "Chloridy", "electrolytes"),
    "P": ("phosphate", "Fosfor", "minerals"),
    "KM": ("uric_acid", "Kyselina mocova", "kidney"),
    "CB": ("total_protein", "Celkova bilkovina", "liver"),
    "ALB": ("albumin", "Albumin", "liver"),
    "PSA": ("psa", "PSA", "hormones"),
    "RBC": ("rbc", "Erytrocyty", "blood_count"),
    "WBC": ("wbc", "Leukocyty", "blood_count"),
    "HCT": ("hematocrit", "Hematokrit", "blood_count"),
    "HGB": ("hemoglobin", "Hemoglobin", "blood_count"),
    "MCV": ("mcv", "MCV", "blood_count"),
    "MCH": ("mch", "MCH", "blood_count"),
    "MCHC": ("mchc", "MCHC", "blood_count"),
    "PLT": ("platelets", "Trombocyty", "blood_count"),
    "RET-ABS": ("reticulocytes_absolute", "Retikulocyty", "blood_count"),
}


FALLBACK_CATEGORY_MAP = {
    "zĂˇkladnĂ­ biochemie": "other",
    "zĂˇkladni biochemie": "other",
    "onkogennĂ­ markery": "other",
    "onkogenni markery": "other",
    "hormony a vitamĂ­ny": "vitamins",
    "hormony a vitaminy": "vitamins",
    "hematologie": "blood_count",
    "minerĂˇly": "minerals",
    "mineraly": "minerals",
}


CSV_HEADERS = [
    "report_date",
    "reported_date",
    "lab_name",
    "fasting_state",
    "marker_key",
    "marker_label",
    "category",
    "value",
    "unit",
    "comparator",
    "reference_low",
    "reference_high",
    "reference_text",
    "status",
    "source_sheet",
    "source_row",
    "notes",
]


STRUCTURAL_LABELS = {
    "Separace séra",
    "Odběr ze žíly",
    "Krevní obraz základní (KO)",
    "Krevní obraz + 5 populační dif.",
    "Separace sĂ©ra",
    "OdbÄ›r ze ĹľĂ­ly",
    "KrevnĂ­ obraz zĂˇkladnĂ­ (KO)",
    "KrevnĂ­ obraz + 5 populaÄŤnĂ­ dif.",
}


@dataclass
class ReviewStats:
    written_rows: int = 0
    skipped_rows: int = 0
    anomalies: list[dict[str, Any]] | None = None
    counts_by_report: Counter | None = None
    counts_by_category: Counter | None = None

    def __post_init__(self) -> None:
        self.anomalies = []
        self.counts_by_report = Counter()
        self.counts_by_category = Counter()


def slugify(text: str) -> str:
    normalized = text.strip().lower()
    normalized = normalized.replace("%", " pct ")
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized)
    return normalized.strip("_")


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def format_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def infer_marker(abbr: str | None, label: str | None, group: str | None) -> tuple[str, str, str]:
    abbr_key = (abbr or "").strip().upper()
    if abbr_key in MARKER_KEY_MAP:
        return MARKER_KEY_MAP[abbr_key]

    label_norm = (label or "").strip().lower()
    if "anti tsh receptor" in label_norm or "trak" in label_norm:
        return ("trak", "TRAK", "thyroid")

    label_text = (label or abbr or "unknown_marker").strip()
    group_key = (group or "").strip().lower()
    category = FALLBACK_CATEGORY_MAP.get(group_key, "other")
    return (slugify(abbr or label_text), label_text, category)


def compute_status(value: float | None, low: float | None, high: float | None) -> str:
    if value is None:
        return "unknown"
    if low is not None and value < low:
        return "low"
    if high is not None and value > high:
        return "high"
    if low is not None or high is not None:
        return "optimal"
    return "unknown"


def build_notes(parts: list[str]) -> str:
    cleaned = [part for part in parts if part]
    return "; ".join(cleaned)


def is_zero_like_numeric(value: Any) -> bool:
    return isinstance(value, (int, float)) and float(value) == 0.0


def normalize_workbook(input_path: Path, output_csv: Path, review_md: Path) -> ReviewStats:
    wb = load_workbook(input_path, data_only=False)
    ws = wb["List 1"]

    stats = ReviewStats()

    output_csv.parent.mkdir(parents=True, exist_ok=True)

    with output_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_HEADERS)
        writer.writeheader()

        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]

            group = row[2].value
            abbr = row[3].value
            unit = row[4].value
            ref_low = parse_float(row[5].value)
            ref_high = parse_float(row[6].value)
            label = row[8].value
            label_text = str(label).strip() if label is not None else ""
            unit_text = str(unit).strip() if unit is not None else ""

            if label_text in STRUCTURAL_LABELS:
                stats.skipped_rows += 1
                stats.anomalies.append(
                    {
                        "row": row_idx,
                        "report": "structural",
                        "marker": label_text,
                        "issue": "non_biomarker_row_skipped",
                        "raw_value": "",
                    }
                )
                continue

            if (is_zero_like_numeric(abbr) or is_zero_like_numeric(unit)) and label_text:
                stats.skipped_rows += 1
                stats.anomalies.append(
                    {
                        "row": row_idx,
                        "report": "source_row",
                        "marker": label_text,
                        "issue": "malformed_marker_metadata_row_skipped",
                        "raw_value": f"abbr={abbr!r}, unit={unit!r}",
                    }
                )
                continue

            marker_key, marker_label, category = infer_marker(
                str(abbr) if abbr is not None else None,
                str(label) if label is not None else None,
                str(group) if group is not None else None,
            )

            for spec in REPORT_SPECS:
                value_cell = row[spec["value_col"] - 1]
                raw_value = value_cell.value
                if raw_value is None or raw_value == "":
                    continue

                notes: list[str] = []
                parsed_value = parse_float(raw_value)

                if value_cell.data_type == "d":
                    stats.skipped_rows += 1
                    stats.anomalies.append(
                        {
                            "row": row_idx,
                            "report": spec["label"],
                            "marker": marker_label,
                            "issue": "date_in_value_column",
                            "raw_value": str(raw_value),
                        }
                    )
                    continue

                if parsed_value is None:
                    stats.skipped_rows += 1
                    stats.anomalies.append(
                        {
                            "row": row_idx,
                            "report": spec["label"],
                            "marker": marker_label,
                            "issue": "non_numeric_value",
                            "raw_value": str(raw_value),
                        }
                    )
                    continue

                reported_date = None
                if spec["date_col"] is not None:
                    reported_date = format_date(row[spec["date_col"] - 1].value)
                    if reported_date is None and spec["fallback_date"]:
                        reported_date = spec["fallback_date"]
                        notes.append(spec["date_note"])
                else:
                    reported_date = spec["fallback_date"]
                    if spec["date_note"]:
                        notes.append(spec["date_note"])

                if reported_date is None:
                    stats.skipped_rows += 1
                    stats.anomalies.append(
                        {
                            "row": row_idx,
                            "report": spec["label"],
                            "marker": marker_label,
                            "issue": "missing_report_date",
                            "raw_value": str(raw_value),
                        }
                    )
                    continue

                normalized_unit = unit_text
                if marker_key == "vitamin_d_25oh" and unit_text.lower() == "mmol/l":
                    normalized_unit = "nmol/L"
                    notes.append("unit_normalized_from_sheet_possible_typo")

                writer.writerow(
                    {
                        "report_date": reported_date,
                        "reported_date": reported_date,
                        "lab_name": spec["label"],
                        "fasting_state": "unknown",
                        "marker_key": marker_key,
                        "marker_label": marker_label,
                        "category": category,
                        "value": parsed_value,
                        "unit": normalized_unit,
                        "comparator": "exact",
                        "reference_low": ref_low if ref_low is not None else "",
                        "reference_high": ref_high if ref_high is not None else "",
                        "reference_text": "",
                        "status": compute_status(parsed_value, ref_low, ref_high),
                        "source_sheet": ws.title,
                        "source_row": row_idx,
                        "notes": build_notes(notes),
                    }
                )
                stats.written_rows += 1
                stats.counts_by_report[spec["label"]] += 1
                stats.counts_by_category[category] += 1

    review_lines = [
        "# Biomarker Intake Review",
        "",
        f"- input: `{input_path}`",
        f"- output: `{output_csv}`",
        f"- written rows: `{stats.written_rows}`",
        f"- skipped rows: `{stats.skipped_rows}`",
        "",
        "## Counts by report",
        "",
    ]
    for label, count in stats.counts_by_report.items():
        review_lines.append(f"- `{label}`: `{count}`")

    review_lines.extend(["", "## Counts by category", ""])
    for label, count in stats.counts_by_category.items():
        review_lines.append(f"- `{label}`: `{count}`")

    review_lines.extend(["", "## Anomalies", ""])
    if not stats.anomalies:
        review_lines.append("- none")
    else:
        for anomaly in stats.anomalies[:20]:
            review_lines.append(
                f"- row `{anomaly['row']}` / `{anomaly['report']}` / `{anomaly['marker']}`: "
                f"`{anomaly['issue']}` (`{anomaly['raw_value']}`)"
            )
        if len(stats.anomalies) > 20:
            review_lines.append(f"- ... and `{len(stats.anomalies) - 20}` more")

    review_md.parent.mkdir(parents=True, exist_ok=True)
    review_md.write_text("\n".join(review_lines) + "\n", encoding="utf-8")
    return stats


def main() -> int:
    if len(sys.argv) != 4:
        print(
            "Usage: normalize-biomarker-workbook.py <input.xlsx> <output.csv> <review.md>",
            file=sys.stderr,
        )
        return 1

    input_path = Path(sys.argv[1])
    output_csv = Path(sys.argv[2])
    review_md = Path(sys.argv[3])

    stats = normalize_workbook(input_path, output_csv, review_md)
    print(
        json.dumps(
            {
                "writtenRows": stats.written_rows,
                "skippedRows": stats.skipped_rows,
                "outputCsv": str(output_csv),
                "reviewMd": str(review_md),
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
